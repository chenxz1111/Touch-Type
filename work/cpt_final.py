import pandas as pd
import xlrd
import openpyxl as op
import re
import docx

def act_word(name, id):
    file = docx.Document(name)
    all_tables = file.tables
    table = all_tables[id]
    para_single = table.cell(1,1).text
    para_single = re.sub(r'[^\u4e00-\u9fa5]','',para_single)
    return len(para_single)

user = ['cxl', 'czf']


work_book = xlrd.open_workbook('time.xlsx')
wb = op.load_workbook("time.xlsx")
sh=wb["Sheet1"]


for id in range(len(user)):
    name = user[id]
    for session in range(3):
        for task in range(5):
            file_name = name + '/' + str(session + 1) + '-' + str(task + 1)
            fPath = 'data/' + file_name + '.docx'
            sh.cell(2 + id + session * 75 + task * 15, 9, act_word(fPath, task))
wb.save("time.xlsx")
