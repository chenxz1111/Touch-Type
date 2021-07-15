import scipy.stats as st
import docx
from data_manager import DataManager
import xlrd
import openpyxl as op
import re
import pandas as pd

def solve(A, B):
    _, p = st.wilcoxon(A, B)
    z = -st.norm.isf(p / 2)
    return z, p

name = ['crj','gtf','yzk','hyh','gjl','czf','cxl','fkz','dyh','lqt','wsj','zyh','cxz','lzj', 'gyz']
speed1 = []
speed2 = []
speed3 = []
accu1 = []
accu2 = []
accu3 = []
tired1 = []
tired2 = []
tired3 = []
worry1 = []
worry2 = []
worry3 = []


for user in name:
    file_name = 'feedback/' + user + '.docx'
    file = docx.Document(file_name)
    all_tables = file.tables
    table = all_tables[0]
    speed1.append(int(table.cell(1,1).text))
    speed2.append(int(table.cell(2,1).text))
    speed3.append(int(table.cell(3,1).text))
    accu1.append(int(table.cell(1,2).text))
    accu2.append(int(table.cell(2,2).text))
    accu3.append(int(table.cell(3,2).text))
    tired1.append(int(table.cell(1,3).text))
    tired2.append(int(table.cell(2,3).text))
    tired3.append(int(table.cell(3,3).text))
    worry1.append(int(table.cell(1,4).text))
    worry2.append(int(table.cell(2,4).text))
    worry3.append(int(table.cell(3,4).text))

s1 = solve(speed1, speed2)
s2 = solve(speed1, speed3)
s3 = solve(speed2, speed3)

a1 = solve(accu1, accu2)
a2 = solve(accu1, accu3)
a3 = solve(accu2, accu3)

t1 = solve(tired1, tired2)
t2 = solve(tired1, tired3)
t3 = solve(tired2, tired3)

w1 = solve(worry1, worry2)
w2 = solve(worry1, worry3)
w3 = solve(worry2, worry3)


work_book = xlrd.open_workbook('zp.xlsx')
wb = op.load_workbook("zp.xlsx")
sh=wb["Sheet1"]

sh.cell(2,2,s1[0])
sh.cell(6,2,s1[1])
sh.cell(3,2,s2[0])
sh.cell(7,2,s2[1])
sh.cell(4,2,s3[0])
sh.cell(8,2,s3[1])

sh.cell(2,3,a1[0])
sh.cell(6,3,a1[1])
sh.cell(3,3,a2[0])
sh.cell(7,3,a2[1])
sh.cell(4,3,a3[0])
sh.cell(8,3,a3[1])

sh.cell(2,4,t1[0])
sh.cell(6,4,t1[1])
sh.cell(3,4,t2[0])
sh.cell(7,4,t2[1])
sh.cell(4,4,t3[0])
sh.cell(8,4,t3[1])

sh.cell(2,5,w1[0])
sh.cell(6,5,w1[1])
sh.cell(3,5,w2[0])
sh.cell(7,5,w2[1])
sh.cell(4,5,w3[0])
sh.cell(8,5,w3[1])

wb.save("feedback.xlsx")

