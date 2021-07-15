import sys
import cv2
import time
import numpy as np
import pickle
from sklearn import svm
from board import Board
import random
import threading
from my_keyboard import MyKeyboard
from frame_data import FrameData
from data_manager import DataManager
import compress_pickle
from train import History
import docx
import os
from shutil import copyfile
import xlrd
import openpyxl as op

class Replay():

    def __init__(self, file_name):
        self.file_name = file_name
        self.init()

    def _run(self):
        self.is_running = True
        while self.is_running:
            if self.auto_play:
                self.incFrame()

    def init(self):
        self.timestamps = pickle.load(open('data/' + self.file_name + '.timestamp', 'rb'))
        self.frames = compress_pickle.load('data/' + self.file_name + '.gz')
        self.frame_id = 0
        self.auto_play = True
        self.history = History()
        self.num = 0
        self.type_time = 0
        self.select_time = 0
        self.delete_time = 0
        self.compt = False
        self.type_start = 0
        self.type_end = 0 
        self.delete_start = 0
        self.delete_end = 0
        self.select_start = 0
        self.select_end = 0
        self.start = 0
        self.end = 0
        self.dtime = False
        self.stime = False
        [self.scalar, self.clf] = pickle.load(open('model/tap.model', 'rb'))

    def incFrame(self):
        if self.frame_id + 1 < len(self.frames):
            frame = self.frames[self.frame_id]
            self.history.updateFrame(frame)
            key_contacts = self.history.getKeyContact(frame)
            for contact in key_contacts:
                if(self.start == 0):
                    timestamp = self.frames[self.frame_id].timestamp
                    self.start = timestamp
                else:
                    timestamp = self.frames[self.frame_id].timestamp
                    self.end = timestamp

                if(iswords(contact.x, contact.y) == True):
                    timestamp = self.frames[self.frame_id].timestamp
                    #self.select_start = timestamp
                    if(self.compt == False):
                        self.type_start = timestamp
                        self.compt = True
                    else:
                        self.type_end = timestamp
                else:
                    if(self.compt == True):
                        if(self.type_end - self.type_start > 0):
                            self.type_time += self.type_end - self.type_start
                            self.type_start = 0
                            self.type_end = 0
                        self.compt = False

                if(isdelete(contact.x, contact.y) == True):
                    timestamp = self.frames[self.frame_id].timestamp
                    self.delete_end = timestamp
                else:
                    timestamp = self.frames[self.frame_id].timestamp
                    if(self.delete_end - self.delete_start > 0):
                        self.delete_time += self.delete_end - self.delete_start
                        self.delete_end = 0
                        self.delete_start = 0
                    self.delete_start = timestamp

                if(isselect(contact.x, contact.y) == True):
                    timestamp = self.frames[self.frame_id].timestamp
                    self.select_end = timestamp
                else:
                    timestamp = self.frames[self.frame_id].timestamp
                    if(self.select_end - self.select_start > 0):
                        self.select_time += self.select_end - self.select_start
                        self.select_end = 0
                        self.select_start = 0
                    self.select_start = timestamp
            self.frame_id += 1
        else:
            self.is_running = False
    
def iswords(x, y):
    if(x > 3.0/23.7 and x < (3.0 + 17.8)/ 23.7 and y > 4.35/13.7 and y < (4.35 + 1.65) / 13.7):
        return True
    if(x > 3.45/23.7 and x < (3.45 + 10 * 1.77)/ 23.7 and y > 6.0/13.7 and y < (6.0 + 1.7) / 13.7):
        return True
    if(x > 4.3/23.7 and x < (4.3 + 9 * 1.78)/ 23.7 and y > 7.7/13.7 and y < (7.7 + 1.65) / 13.7):
        return True
    return False

def isselect(x, y):
    if(x > 2.3/23.7 and x < (2.3 + 17.5)/ 23.7 and y > 2.6/13.7 and y < (2.6 + 1.75) / 13.7):
        return True
    if(x > 7.79/23.7 and x < (7.79 + 8.8)/ 23.7 and y > 9.35/13.7 and y < (9.35 + 2.05) / 13.7):
        return True
    return False

def isdelete(x, y):
    if(x > 20.8/23.7 and x < (20.8 + 2.5)/ 23.7 and y > 4.35/13.7 and y < (4.35 + 1.65) / 13.7):
        return True
    return False


if __name__ == "__main__":
    board = 3
    keyboard = MyKeyboard()
    users = ['cxl','czf','dhy','hyh','crj','yzk','gtf','fkz','gyz','wsj','zyh','gjl','cxz','lzj','lqt']
    for u in range(len(users)):
        print(users[u])
        for session in range(5):
            file_name = users[u] + '/'+ str(board) +'-' + str(session + 1)
            replay = Replay(file_name)
            replay._run()
            totol_time = replay.end - replay.start
            others =  totol_time - replay.type_time - replay.select_time - replay.delete_time
    # print('type_time:',replay.type_time)
    # print('select_time:',replay.select_time)
    # print('delete_time:',replay.delete_time)
    # print('others:', others)
            work_book = xlrd.open_workbook('time_new.xlsx')
            wb = op.load_workbook("time_new.xlsx")
            sh=wb["Sheet1"]
            n = board
            m = session + 1
    # print(n, m)
            id = u + 2
            sh.cell(id + (n - 1) * 75+(m -1) * 15, 4, replay.type_time)
            sh.cell(id + (n - 1) * 75+(m -1) * 15, 5, replay.select_time)
            sh.cell(id + (n - 1) * 75+(m -1) * 15, 6, replay.delete_time)
            sh.cell(id + (n - 1) * 75+(m -1) * 15, 7, others)
            wb.save("time_new.xlsx")