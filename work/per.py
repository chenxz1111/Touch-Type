import sys
import cv2
import time
import numpy as np
import pickle
from sklearn import svm
from board import Board
import random
import threading
from my_keyboard import MyKeyboard
from frame_data import FrameData
from data_manager import DataManager
import compress_pickle
from train import History
import docx
import os
from shutil import copyfile
import xlrd
import openpyxl as op

class Replay():

    def __init__(self, file_name):
        self.file_name = file_name
        self.init()

    def _run(self):
        self.is_running = True
        while self.is_running:
            if self.auto_play:
                self.incFrame()

    def init(self):
        self.timestamps = pickle.load(open('data/' + self.file_name + '.timestamp', 'rb'))
        self.frames = compress_pickle.load('data/' + self.file_name + '.gz')
        self.frame_id = 0
        self.auto_play = True
        self.history = History()
        self.nnum = 0
        self.snum  =0 
        [self.scalar, self.clf] = pickle.load(open('model/tap.model', 'rb'))

    def incFrame(self):
        if self.frame_id + 1 < len(self.frames):
            frame = self.frames[self.frame_id]
            self.history.updateFrame(frame)
            key_contacts = self.history.getKeyContact(frame)
            for contact in key_contacts:
                if(isn(contact.x, contact.y)):
                    self.nnum += 1
                if(iss(contact.x, contact.y)):
                    self.snum += 1
            self.frame_id += 1
        else:
            self.is_running = False
    
def isn(x, y):
    if(x > (2.3 + 1.75)/23.7 and x < (2.3 + 1.75 * 7)/ 23.7 and y > 2.6/13.7 and y < (2.6 + 1.75) / 13.7):
        return True
    return False

def iss(x, y):
    if(x > 7.79/23.7 and x < (7.79 + 8.8)/ 23.7 and y > 9.35/13.7 and y < (9.35 + 2.05) / 13.7):
        return True
    if(x > 2.3/23.7 and x < (2.3 + 1.75)/ 23.7 and y > 2.6/13.7 and y < (2.6 + 1.75) / 13.7):
        return True
    return False


if __name__ == "__main__":
    for b in range(3):
        board = b + 1
        keyboard = MyKeyboard()
        users = ['cxl','czf','dhy','hyh','crj','yzk','gtf','fkz','gyz','wsj','zyh','gjl','cxz','lzj','lqt']
        for u in range(len(users)):
            print(users[u])
            for session in range(5):
                file_name = users[u] + '/'+ str(board) +'-' + str(session + 1)
                replay = Replay(file_name)
                replay._run()
                # sp = replay.snum / (replay.snum + replay.nnum)
                # np = replay.nnum / (replay.snum + replay.nnum)
        # print('type_time:',replay.type_time)
        # print('select_time:',replay.select_time)
        # print('delete_time:',replay.delete_time)
        # print('others:', others)
                work_book = xlrd.open_workbook('percentage_new.xlsx')
                wb = op.load_workbook("percentage_new.xlsx")
                sh=wb["Sheet1"]
                n = board
                m = session + 1
        # print(n, m)
                id = u + 2
                sh.cell(id + (n - 1) * 75+(m -1) * 15, 6, replay.snum)
                sh.cell(id + (n - 1) * 75+(m -1) * 15, 7, replay.nnum)
                wb.save("percentage_new.xlsx")