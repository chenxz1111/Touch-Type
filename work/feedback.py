import scipy.stats as st
import docx
from data_manager import DataManager
import xlrd
import openpyxl as op
import re
import pandas as pd

name = ['crj','gtf','yzk','hyh','gjl','czf','cxl','fkz','dyh','lqt','wsj','zyh','cxz','lzj', 'gyz']


work_book = xlrd.open_workbook('feedback.xlsx')
wb = op.load_workbook("feedback.xlsx")
sh=wb["Sheet1"]

id = 2

for user in name:
    file_name = 'feedback/' + user + '.docx'
    file = docx.Document(file_name)
    all_tables = file.tables
    table = all_tables[0]
    sh.cell(id,1,user)
    sh.cell(id,2,'1')
    sh.cell(id,3,int(table.cell(1,1).text))
    sh.cell(id,4,int(table.cell(1,2).text))
    sh.cell(id,5,int(table.cell(1,3).text))
    sh.cell(id,6,int(table.cell(1,4).text))
    id += 1
    sh.cell(id,1,user)
    sh.cell(id,2,2)
    sh.cell(id,3,int(table.cell(2,1).text))
    sh.cell(id,4,int(table.cell(2,2).text))
    sh.cell(id,5,int(table.cell(2,3).text))
    sh.cell(id,6,int(table.cell(2,4).text))
    id += 1
    sh.cell(id,1,user)
    sh.cell(id,2,3)
    sh.cell(id,3,int(table.cell(3,1).text))
    sh.cell(id,4,int(table.cell(3,2).text))
    sh.cell(id,5,int(table.cell(3,3).text))
    sh.cell(id,6,int(table.cell(3,4).text))
    id += 1

wb.save("feedback.xlsx")

