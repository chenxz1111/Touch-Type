import sys
import cv2
import time
import numpy as np
import pickle
from sklearn import svm
from board import Board
import random
import threading
from my_keyboard import MyKeyboard
from frame_data import FrameData
from data_manager import DataManager
import compress_pickle
from train import History
import docx
import os
from shutil import copyfile
import xlrd
import openpyxl as op

class Replay():

    def __init__(self, file_name):
        self.file_name = file_name
        self.init()

    def _run(self):
        self.is_running = True
        while self.is_running:
            if self.auto_play:
                self.incFrame()

    def init(self):
        self.timestamps = pickle.load(open('data/' + self.file_name + '.timestamp', 'rb'))
        self.frames = compress_pickle.load('data/' + self.file_name + '.gz')
        self.frame_id = 0
        self.auto_play = True
        self.history = History()
        self.delete = True
        self.point = []
        [self.scalar, self.clf] = pickle.load(open('model/tap.model', 'rb'))

    def incFrame(self):
        if self.frame_id + 1 < len(self.frames):
            frame = self.frames[self.frame_id]
            self.history.updateFrame(frame)
            key_contacts = self.history.getKeyContact(frame)
            for contact in key_contacts:
                if(iswords(contact.x, contact.y)[0] == True):
                    self.delete = True
                    self.point.append((iswords(contact.x, contact.y)[1], act_value(contact.x, contact.y)[0], act_value(contact.x, contact.y)[1]))
                elif(isselect(contact.x, contact.y) == True):
                    self.delete = False
                elif(isdelete(contact.x, contact.y) == True):
                    if(self.delete == True):
                        if(len(self.point) > 0):
                            self.point.pop()
            self.frame_id += 1
        else:
            self.is_running = False
    
def iswords(x, y):
    qline = ['q', 'w', 'e', 'r', 't', 'y', 'u', 'i', 'o', 'p']
    aline = ['a', 's', 'd', 'f', 'g', 'h', 'j', 'k', 'l']
    zline = ['z', 'x', 'c', 'v', 'b', 'n', 'm']
    if(x > 3.0/23.7 and x < (3.0 + 17.8)/ 23.7 and y > 4.35/13.7 and y < (4.35 + 1.65) / 13.7):
        id = int((x * 23.7 - 3.0)/1.78)
        return True, qline[id]
    if(x > 3.45/23.7 and x < (3.45 + 9 * 1.77)/ 23.7 and y > 6.0/13.7 and y < (6.0 + 1.7) / 13.7):
        id = int((x * 23.7 - 3.45)/1.77)
        return True, aline[id]
    if(x > 4.3/23.7 and x < (4.3 + 7 * 1.78)/ 23.7 and y > 7.7/13.7 and y < (7.7 + 1.65) / 13.7):
        id = int((x * 23.7 - 4.3)/1.78)
        return True, zline[id]
    return False, '_'

def act_value(x, y):
    return x *23.7, y * 13.7

def isselect(x, y):
    if(x > 2.3/23.7 and x < (2.3 + 17.5)/ 23.7 and y > 2.6/13.7 and y < (2.6 + 1.75) / 13.7):
        return True
    if(x > 7.79/23.7 and x < (7.79 + 8.8)/ 23.7 and y > 9.35/13.7 and y < (9.35 + 2.05) / 13.7):
        return True
    return False

def isdelete(x, y):
    if(x > 20.8/23.7 and x < (20.8 + 2.5)/ 23.7 and y > 4.35/13.7 and y < (4.35 + 1.65) / 13.7):
        return True
    return False


if __name__ == "__main__":
    
    keyboard = MyKeyboard()
    users = ['cxl','czf','dhy','hyh','crj','yzk','gtf','fkz','gyz','wsj','zyh','gjl','cxz','lzj','lqt']
    for user in users:
        print(user)
        for session in range(5):
            file_name = user + '/3-' + str(session + 1)
            replay = Replay(file_name)
            replay._run()
            
            #print(replay.point)
            work_book = xlrd.open_workbook('pointset.xlsx')
            wb = op.load_workbook("pointset.xlsx")
            sh=wb["Sheet1"]
            sheet_1 = work_book.sheet_by_index(0)
            row_sum = sheet_1.nrows
            for i in range(len(replay.point)):
                row_sum += 1
                sh.cell(row_sum , 1, user)
                sh.cell(row_sum , 2, 3)
                sh.cell(row_sum , 3, replay.point[i][0])
                sh.cell(row_sum , 4, replay.point[i][1])
                sh.cell(row_sum , 5, replay.point[i][2])
            wb.save("pointset.xlsx")